from flask import Flask, render_template, request, jsonify, Response
import cv2
import numpy as np
import os
import json
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment
from io import BytesIO

app = Flask(__name__, template_folder='templates', static_folder='static')
classifier = 'haarcascade_frontalface_default.xml'
image_dir = 'images'
haar_cascade = cv2.CascadeClassifier(classifier)
webcam = cv2.VideoCapture(0)
model = None
names = {}
is_running = False

def load_training_data():
    global model, names
    (images, labels, names, id) = ([], [], {}, 0)
    for (subdirs, dirs, files) in os.walk(image_dir):
        for subdir in dirs:
            names[id] = subdir
            subjectpath = os.path.join(image_dir, subdir)
            for filename in os.listdir(subjectpath):
                if not filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.pgm')):
                    continue
                path = os.path.join(subjectpath, filename)
                images.append(cv2.imread(path, 0))
                labels.append(id)
            id += 1
    if images:
        (images, labels) = [np.array(lis) for lis in [images, labels]]
        model = cv2.face.LBPHFaceRecognizer_create()
        model.train(images, labels)

load_training_data()

def log_attendance(name, date, time, photo_path):
    log = {}
    if os.path.exists('attendance_log.json'):
        with open('attendance_log.json', 'r') as f:
            log = json.load(f)
    if date not in log:
        log[date] = []
    log[date].append({"name": name, "time": time, "photo": photo_path})
    with open('attendance_log.json', 'w') as f:
        json.dump(log, f, indent=4)

def gen_frames(mode='attendance'):
    global is_running
    size = 2
    (im_width, im_height) = (112, 92)
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    logged_faces = set()
    
    if mode == 'training':
        name = app.config.get('training_name', '')
        path = os.path.join(image_dir, name)
        os.makedirs(path, exist_ok=True)
        pin = sorted([int(n[:n.find('.')]) for n in os.listdir(path) if n.endswith('.png')] + [0])[-1] + 1
        count = 0
        count_max = 20

    while is_running:
        rval, frame = webcam.read()
        if not rval:
            continue
        
        frame = cv2.flip(frame, 1)
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        mini = cv2.resize(gray, (int(gray.shape[1] / size), int(gray.shape[0] / size)))
        faces = haar_cascade.detectMultiScale(mini)

        for face_i in faces:
            (x, y, w, h) = [v * size for v in face_i]
            face = gray[y:y + h, x:x + w]
            face_resize = cv2.resize(face, (im_width, im_height))
            
            if mode == 'attendance' and model:
                prediction = model.predict(face_resize)
                name = names[prediction[0]] if prediction[1] < 90 else "Unknown"
                if name != "Unknown" and name not in logged_faces:
                    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
                    photo_filename = f"{name}_{today}_{timestamp.replace(':', '-')}.png"
                    os.makedirs('logIMG', exist_ok=True)
                    photo_path = os.path.join('logIMG', photo_filename)
                    cv2.imwrite(photo_path, face_resize)
                    log_attendance(name, today, timestamp, photo_path)
                    logged_faces.add(name)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, f"{name} - {prediction[1]:.0f}", (x+5, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
                      
            elif mode == 'training' and count < count_max:
                cv2.imwrite(f'{path}/{pin}.png', face_resize)
                pin += 1
                count += 1
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, f"Training: {count}/{count_max}", (x+5, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)

        ret, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/')
def index():
    log = {}
    if os.path.exists('attendance_log.json'):
        with open('attendance_log.json', 'r') as f:
            log = json.load(f)
    return render_template('index.html', log=log)

@app.route('/start_training', methods=['POST'])
def start_training():
    global is_running
    name = request.form.get('name')
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    app.config['training_name'] = name
    is_running = True
    return jsonify({'message': 'Training started'})

@app.route('/start_attendance', methods=['POST'])
def start_attendance():
    global is_running
    is_running = True
    return jsonify({'message': 'Attendance started'})

@app.route('/stop', methods=['POST'])
def stop():
    global is_running
    is_running = False
    load_training_data()
    return jsonify({'message': 'Stopped'})

@app.route('/video_feed/<mode>')
def video_feed(mode):
    return Response(gen_frames(mode), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/export_csv/<date>', methods=['GET'])
def export_csv(date):
    log = {}
    if os.path.exists('attendance_log.json'):
        with open('attendance_log.json', 'r') as f:
            log = json.load(f)
    
    if date not in log:
        return jsonify({'error': 'No records for this date'}), 404
    
    # Filter earliest record per person
    earliest_records = {}
    for record in log[date]:
        name = record['name']
        time = record['time']
        if name not in earliest_records or time < earliest_records[name]['time']:
            earliest_records[name] = record
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {date}"
    
    # Define headers
    headers = ['Date', 'Name', 'Time', 'Photo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
    
    # Define border style (0.5pt)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Add data and images
    for row, record in enumerate(earliest_records.values(), 2):
        # Write data
        ws.cell(row=row, column=1).value = date
        ws.cell(row=row, column=2).value = record['name']
        ws.cell(row=row, column=3).value = record['time']
        
        # Embed image
        photo_path = record['photo']
        if os.path.exists(photo_path):
            img = Image(photo_path)
            img.width = 50  # Adjust image size
            img.height = 50
            ws.add_image(img, f'D{row}')
        
        # Apply border and alignment to cells
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', indent=1)  # Add padding via indent
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Date
    ws.column_dimensions['B'].width = 15  # Name
    ws.column_dimensions['C'].width = 15  # Time
    ws.column_dimensions['D'].width = 15  # Photo
    
    # Adjust row heights to accommodate images
    for row in range(2, len(earliest_records) + 2):
        ws.row_dimensions[row].height = 60  # Adjust for image height
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment;filename=attendance_{date}.xlsx'}
    )

from flask import send_from_directory

@app.route('/<path:filename>')
def serve_root_file(filename):
    return send_from_directory('.', filename)

if __name__ == '__main__':
    app.run(debug=True, host='localhost', port=5000)